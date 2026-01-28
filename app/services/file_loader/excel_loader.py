from __future__ import annotations

from pathlib import Path
from typing import List

from app.core.exceptions import ServiceException
from app.services.file_loader.base import FileLoader, PageContent


def _parse_xlsx(file_path: Path) -> List[PageContent]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ServiceException("缺少 openpyxl 依赖，无法解析 xlsx") from exc

    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    pages: List[PageContent] = []

    for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
        # 每个工作表视作一页
        sheet = workbook[sheet_name]
        rows: List[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        text = "\n".join(rows).strip()
        pages.append(PageContent(page_number=sheet_index, page_label=sheet_name, text=text))

    return pages


def _parse_xls(file_path: Path) -> List[PageContent]:
    try:
        import xlrd
    except Exception as exc:
        raise ServiceException("缺少 xlrd 依赖，无法解析 xls") from exc

    workbook = xlrd.open_workbook(str(file_path))
    pages: List[PageContent] = []

    for sheet_index in range(workbook.nsheets):
        # 每个工作表视作一页
        sheet = workbook.sheet_by_index(sheet_index)
        rows: List[str] = []
        for row_index in range(sheet.nrows):
            values = [str(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        text = "\n".join(rows).strip()
        pages.append(PageContent(page_number=sheet_index + 1, page_label=sheet.name, text=text))

    return pages


class ExcelLoader(FileLoader):
    """Excel 解析器，支持 xlsx 及 xls。"""

    def parse(self, file_path: Path) -> List[PageContent]:
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return _parse_xlsx(file_path)
        if suffix == ".xls":
            return _parse_xls(file_path)
        raise ServiceException(f"不支持的 Excel 格式: {suffix}")
