from __future__ import annotations

from pathlib import Path
from typing import List, Optional

from app.exception.exceptions import ServiceException
from app.rag.file_loader.base import FileLoader, PageContent


def _parse_xlsx(file_path: Path) -> List[PageContent]:
    """
    解析 .xlsx 文件：每个工作表视为一页。

    Args:
        file_path: .xlsx 文件路径

    Returns:
        按工作表组织的页面内容列表

    Raises:
        ServiceException: 缺少 openpyxl 依赖
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ServiceException("缺少 openpyxl 依赖，无法解析 xlsx") from exc

    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    pages: List[PageContent] = []

    for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
        # 每个工作表视作一页
        sheet = workbook[sheet_name]
        rows: List[str] = []
        for row in sheet.iter_rows(values_only=True):
            # 将单元格值转为字符串，用制表符分隔
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        text = "\n".join(rows).strip()
        pages.append(
            PageContent(page_number=sheet_index, page_label=sheet_name, text=text)
        )

    return pages


def _parse_xls(file_path: Path) -> List[PageContent]:
    """
    解析 .xls 文件：每个工作表视为一页。

    Args:
        file_path: .xls 文件路径

    Returns:
        按工作表组织的页面内容列表

    Raises:
        ServiceException: 缺少 xlrd 依赖
    """
    try:
        import xlrd
    except Exception as exc:
        raise ServiceException("缺少 xlrd 依赖，无法解析 xls") from exc

    workbook = xlrd.open_workbook(str(file_path))
    pages: List[PageContent] = []

    for sheet_index in range(workbook.nsheets):
        # 每个工作表视作一页
        sheet = workbook.sheet_by_index(sheet_index)
        rows: List[str] = []
        for row_index in range(sheet.nrows):
            # 将单元格值转为字符串，用制表符分隔
            values = [
                str(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)
            ]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        text = "\n".join(rows).strip()
        pages.append(
            PageContent(page_number=sheet_index + 1, page_label=sheet.name, text=text)
        )

    return pages


class ExcelLoader(FileLoader):
    """Excel 解析器，支持 xlsx 及 xls。"""

    def parse(
            self, file_path: Path, output_dir: Optional[Path] = None
    ) -> List[PageContent]:
        """
        解析 Excel 文件。

        Args:
            file_path: Excel 文件路径
            output_dir: 未使用（Excel 解析无图片提取）

        Returns:
            按工作表组织的页面内容列表

        Raises:
            ServiceException: 不支持的文件格式
        """
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return _parse_xlsx(file_path)
        if suffix == ".xls":
            return _parse_xls(file_path)
        raise ServiceException(f"不支持的 Excel 格式: {suffix}")
