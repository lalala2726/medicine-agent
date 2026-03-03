from __future__ import annotations

from pathlib import Path

from app.core.exception.exceptions import ServiceException
from app.rag.file_loader.parsers.base import BaseParser
from app.rag.file_loader.types import ParsedPage


def _parse_xlsx(file_path: Path) -> list[ParsedPage]:
    """
    功能描述:
        解析 xlsx 文件，每个工作表视作一页，列值使用制表符连接。

    参数说明:
        file_path (Path): xlsx 文件路径。

    返回值:
        list[ParsedPage]: 按工作表组织的页面列表。

    异常说明:
        ServiceException: 缺少 openpyxl 依赖时抛出。
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ServiceException("缺少 openpyxl 依赖，无法解析 xlsx 文件") from exc

    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    pages: list[ParsedPage] = []
    for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[sheet_name]
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        pages.append(
            ParsedPage(
                page_number=sheet_index,
                page_label=sheet_name,
                text="\n".join(rows),
            )
        )
    return pages


def _parse_xls(file_path: Path) -> list[ParsedPage]:
    """
    功能描述:
        解析 xls 文件，每个工作表视作一页，列值使用制表符连接。

    参数说明:
        file_path (Path): xls 文件路径。

    返回值:
        list[ParsedPage]: 按工作表组织的页面列表。

    异常说明:
        ServiceException: 缺少 xlrd 依赖时抛出。
    """
    try:
        import xlrd
    except Exception as exc:
        raise ServiceException("缺少 xlrd 依赖，无法解析 xls 文件") from exc

    workbook = xlrd.open_workbook(str(file_path))
    pages: list[ParsedPage] = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        rows: list[str] = []
        for row_index in range(sheet.nrows):
            values = [
                str(sheet.cell_value(row_index, col_index))
                for col_index in range(sheet.ncols)
            ]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        pages.append(
            ParsedPage(
                page_number=sheet_index + 1,
                page_label=sheet.name,
                text="\n".join(rows),
            )
        )
    return pages


class ExcelParser(BaseParser):
    """
    功能描述:
        解析 Excel 文件，支持 xlsx 与 xls。

    参数说明:
        无。解析参数通过 `parse` 方法传入。

    返回值:
        无。调用 `parse` 时返回页面列表。

    异常说明:
        ServiceException: 文件后缀不支持或依赖缺失时抛出。
    """

    def parse(self, file_path: Path) -> list[ParsedPage]:
        """
        功能描述:
            根据后缀分发 Excel 解析分支。

        参数说明:
            file_path (Path): Excel 文件路径。

        返回值:
            list[ParsedPage]: 按工作表组织的页面列表。

        异常说明:
            ServiceException: 不支持的 Excel 文件格式时抛出。
        """
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return _parse_xlsx(file_path)
        if suffix == ".xls":
            return _parse_xls(file_path)
        raise ServiceException(f"不支持的 Excel 格式: {suffix}")
